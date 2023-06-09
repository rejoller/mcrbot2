import telebot
from google.oauth2 import service_account
from googleapiclient.discovery import build
from telebot import types

import json
import time
from urllib3.exceptions import ReadTimeoutError
from fuzzywuzzy import fuzz
from prettytable import PrettyTable
from pretty_html_table import build_table
import openpyxl
import logging
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import os
import matplotlib.pyplot as plt
import numpy as np
import tempfile
import pandas as pd
import gspread

user_messages = {}
# Добавьте этот словарь в начало вашего кода
additional_info_storage = {}
espd_info_storage = {}
szoreg_info_storage = {}
message_storage = {}
districts = ["Абанский р-н", "Ачинский р-н", "Курагинский р-н"]

# Установка токена и создание бота
bot_token = '6263941409:AAE20_qJIMTw03dBYoH0_xcbugDs_4FzA5Y'
bot = telebot.TeleBot(bot_token)


import csv
from datetime import datetime

def log_user_data(user_id, first_name, last_name, username, message_text):
    file_path = 'users_data.csv'
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Проверяем, существует ли файл. Если нет, создаем его с заголовками
    try:
        with open(file_path, 'x', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Timestamp', 'User ID', 'First Name', 'Last Name', 'Username', 'Message'])
    except FileExistsError:
        pass

    # Записываем данные пользователя в файл
    with open(file_path, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([current_time, user_id, first_name, last_name, username, message_text])

def log_user_data_from_message(message):
    user_id = message.from_user.id
    first_name = message.from_user.first_name
    last_name = message.from_user.last_name
    username = message.from_user.username
    message_text = message.text

    log_user_data(user_id, first_name, last_name, username, message_text)



@bot.message_handler(commands=['help'])
def handle_help_command(message):
    log_user_data_from_message(message)
    help_text = (
        'Введи название населенного пункта или муниципального образования, чтобы получить информацию о связи. Чтобы узнать информацию о сотовой связи, выбери /2g /3g или /4g. Чтобы получить информацию о населенных пунктах без сотовой связи жми /nomobile\n\n'
        'Для получения списка ФАП из контракта с ПАО "Ростелеком" нажми /fp\n'
        'Для получения списка точек Аг.ГОиЧС из контракта с ПАО "Ростелеком" нажми /ago\n\n'
        'Чтобы узнать о подключении к ТОРКНД, введи сообщение "тор" и наименование муниципального образования. '
        'Например, "тор Енисейский".\n'
        'Если нужна статистика по всему краю, жми /knd_kraj\n\n'
        'Чтобы узнать, кто сегодня в отпуске, жми /otpusk\n\n'
        'Если остались вопросы, пиши @rejoller.')
    bot.send_message(message.chat.id, help_text)


@bot.message_handler(commands=['knd_kraj'])
def handle_knd_kraj_command(message):
    log_user_data_from_message(message)
    bot.send_message(message.chat.id, 'Загружаю данные')
    handle_knd_kraj_message(message)

@bot.message_handler(commands=['fp'])
def handle_fp_command(message):
    # Загрузите данные из файла с информацией о населенных пунктах
    user_first_name = message.from_user.first_name
    bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌')
    log_user_data_from_message(message)
    fp_data, fp_headers = load_fp_data()

    # Передайте fp_data и fp_headers в функцию handler_fp_message
    handler_fp_message(message, fp_data, fp_headers)
    del fp_data
    del fp_headers

@bot.message_handler(commands=['ago'])
def handle_ago_command(message):
    # Загрузите данные из Google Sheets с информацией об АгГОиЧС
    user_first_name = message.from_user.first_name
    bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌')
    log_user_data_from_message(message)
    aggoics_data, aggoics_headers = load_aggoics_data()

    # Передайте aggoics_data и aggoics_headers в функцию handler_aggoics_message
    handler_aggoics_message(message, aggoics_data, aggoics_headers)
    del aggoics_data
    del aggoics_headers

from datetime import datetime, timedelta

@bot.message_handler(commands=['otpusk'])
def handle_otpusk_command(message, days_ahead=14):
    # Загрузите данные из файла с информацией об отпусках
    bot.send_message(message.chat.id, 'Загружаю данные')
    log_user_data_from_message(message)
    otpusk_data = load_otpusk_data()

    # Получите список сотрудников, которые сегодня в отпуске и уходят в отпуск в ближайшие 3 дня
    employees_on_vacation, employees_starting_vacation_soon = get_employees_on_vacation(otpusk_data, days_ahead)

    response = ""

    if employees_on_vacation:
        response += 'Сегодня в отпуске:\n\n'
        for row in employees_on_vacation:
            response += f"{row[0]}, {row[1]}\n"
            response += f"Дата начала отпуска: {row[3]}\n"
            response += f"Дата окончания отпуска: {row[4]}\n\n"

    if employees_starting_vacation_soon:
        response += f"\nСотрудники, уходящие в отпуск в ближайшие {days_ahead} дней:\n\n"
        for row in employees_starting_vacation_soon:
            response += f"{row[0]}, {row[1]}\n"
            response += f"Дата начала отпуска: {row[3]}\n"
            response += f"Дата окончания отпуска: {row[4]}\n\n"

    if not response:
        response = "Сегодня никто не в отпуске, и никто не уходит в отпуск в ближайшие 14 дней."

    bot.send_message(message.chat.id, response)

# Установка параметров доступа к API Google Sheets
SCOPES = ['https://www.googleapis.com/auth/spreadsheets.readonly']
SERVICE_ACCOUNT_FILE = '/home/rejoller/mcrbot/credentials.json'
creds = None
creds = service_account.Credentials.from_service_account_file(
    SERVICE_ACCOUNT_FILE, scopes=SCOPES)

# ID таблицы
SPREADSHEET_ID = '1lA6wXSOmi6nj4HDOpFdzm4_KaUQAAakNNxOyXx7p1ZQ'


def split_message(message, max_length=4096):
    if len(message) <= max_length:
        return [message]

    messages = []
    while len(message) > max_length:
        split_index = message[:max_length].rfind('\n')
        if split_index == -1:
            split_index = max_length

        messages.append(message[:split_index])
        message = message[split_index:].lstrip()

    if message:
        messages.append(message)

    return messages


def split_message_table(headers, data, max_message_length=4096):
    table = PrettyTable()
    table.field_names = headers

    for row in data:
        if len(row) == len(headers):
            table.add_row(row)
        else:
            print(f"Skipping row with incorrect number of values: {row}")

    table_str = table.get_string()

    messages = []
    lines = table_str.split('\n')
    current_message = lines[0] + '\n' + lines[1] + '\n'

    for row in lines[2:]:
        test_message = current_message + row + '\n'

        if len(test_message) <= max_message_length:
            current_message = test_message
        else:
            messages.append(f"<pre>{current_message.strip()}</pre>")
            current_message = lines[0] + '\n' + lines[1] + '\n' + row + '\n'

    if current_message:
        messages.append(f"<pre>{current_message.strip()}</pre>")

    return messages



def get_value(row, index, default_value=''):
    try:
        return row[index]
    except IndexError:
        return default_value


def normalize_text_v2(text):
    text = text.lower().replace('ё', 'е').replace('р-н', 'район').replace('-', ' ')
    text = re.sub(r'(N|№|No)', 'N', text, flags=re.IGNORECASE)
    text = text.replace(' район', '').strip()
    return text


def search_values(query):
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='goroda2.0!A1:T1721').execute()
    rows = result.get('values', [])
    normalized_query = normalize_text_v2(query)
    found_values_a = [row for row in rows if normalized_query == normalize_text_v2(row[0])]
    found_values_s = [row for row in rows if fuzz.token_sort_ratio(normalized_query, normalize_text_v2(row[18])) >= 99]

    return found_values_a, found_values_s


headers = ['Наименование', 'Население', 'Сотовая связь', 'Интернет', 'Программа', 'Таксофон']



from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

def create_excel_file(headers, data):
    wb = Workbook()
    ws = wb.active

    # Шрифт и выравнивание заголовков
    header_font = Font(name='Arial', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Шрифт и выравнивание данных
    data_font = Font(name='Arial')
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Границы ячеек
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Заливка фона для заголовков
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        cell.fill = header_fill

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_data)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    file_name = "test_file.xlsx"
    wb.save(file_name)
    print(f"File saved as {file_name}")
    return file_name

def create_excel_file_2(headers, data):
    wb = Workbook()
    ws = wb.active

    # Шрифт и выравнивание заголовков
    header_font = Font(name='Arial', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Шрифт и выравнивание данных
    data_font = Font(name='Arial')
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # Границы ячеек
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Заливка фона для заголовков
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
        cell.fill = header_fill

    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_data in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_data)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


import io
from openpyxl.styles import Font, Alignment, Border, Side



def adjust_column_width(worksheet):
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = str(cell.value)
            length = len(value.encode('utf-8'))
            if length > max_length:
                max_length = length

        # Настройка ширины столбца
        estimated_width = max_length * 0.7  # Умножение на коэффициент для учета разных ширин символов
        worksheet.column_dimensions[column].width = estimated_width



from io import BytesIO

def convert_to_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    # Настраиваем стили для заголовков
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="AED6F1",
                              end_color="AED6F1",
                              fill_type="solid")

    # Настраиваем стили для данных
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    data_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    data_fill = PatternFill(start_color="ECECEC",
                            end_color="ECECEC",
                            fill_type="solid")

    for row_idx, row in enumerate(data, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Применяем стили
            if row_idx == 1:
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                cell.fill = header_fill
            else:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = data_border
                if row_idx % 2 == 0:
                    cell.fill = data_fill

    # Вызов функции для автоматической настройки ширины столбцов
    adjust_column_width(ws)

    # Добавляем автофильтр
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    # Закрепляем строку заголовка
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def load_goroda_data():
    # Загружаем данные из Google Sheets
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='goroda2.0!A1:T1721').execute()
    rows = result.get('values', [])
    return rows

import folium
from telebot.types import InputFile
from folium.plugins import MarkerCluster


def create_map_with_markers(rows):
    map_with_markers = folium.Map(location=[59.664482, 91.913147], zoom_start=10)

    # Создаем кластер маркеров
    marker_cluster = MarkerCluster().add_to(map_with_markers)

    for row in rows:
        if row[7] and row[8]:  # проверяем, есть ли широта и долгота
            folium.Marker(
                location=[float(row[7]), float(row[8])],
                popup=row[1],
                icon=None,
            ).add_to(marker_cluster)

    # Добавляем встроенный стиль для скрытия элемента с классом leaflet-control-attribution
    map_with_markers.get_root().html.add_child(folium.Element("<style>.leaflet-control-attribution { display: none; }</style>"))

    return map_with_markers


def webAppKeyboard(url):
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    webAppTest = types.WebAppInfo(url)
    one_butt = types.InlineKeyboardButton(text="для мобильных устройств", web_app=webAppTest)
    two_butt = types.InlineKeyboardButton(text="ПК", url=url)
    keyboard.add(one_butt, two_butt)
    return keyboard #возвращаем клавиатуру


def webAppKeyboard_jt(url):
    keyboard = types.InlineKeyboardMarkup(row_width=1)
    webAppTest = types.WebAppInfo(url)
    button = types.InlineKeyboardButton(text="открыть секретную страницу", web_app=webAppTest)

    keyboard.add(button)
    return keyboard #возвращаем клавиатуру




def filter_and_send_data(message, filter_func, command):
    goroda_data = load_goroda_data()
    headers = ['Наименование населенного пункта', 'Население 2010', 'Население 2020', 'Сотовая связь', 'Программа']
    filtered_data = [headers]
    filtered_goroda_data = []

    for row in goroda_data:
        if filter_func(row):
            filtered_row = [row[i] if i < len(row) else '' for i in [1, 2, 5, 3, 11]]
            filtered_data.append(filtered_row)
            filtered_goroda_data.append(row)

    # Создать карту с маркерами
    map_with_markers = create_map_with_markers(filtered_goroda_data)
    map_filename = f"{command}_map.html"
    map_with_markers.save(map_filename)

    # Конвертировать данные в формат Excel и отправить
    buffer = convert_to_excel(filtered_data)
    filename = f"{command}.xlsx"
    with open(filename, "wb") as excel_file:
        excel_file.write(buffer.getvalue())

    with open(filename, "rb") as excel_file:
        document = InputFile(excel_file)
        bot.send_document(message.chat.id, document=document, caption="Список населенных пунктов")

    os.remove(filename)

    # Отправить файл с картой
   # with open(map_filename, "rb") as map_file:
    #    document = InputFile(map_file)
      #  bot.send_document(message.chat.id, document=document, caption=map_filename)

    os.remove(map_filename)
    url = f"https://rejoller.pythonanywhere.com/{command}"
    bot.send_message(message.chat.id, "Чтобы посмотреть карту, нажмите кнопку ниже", reply_markup=webAppKeyboard(url))




def filter_2g(row):
    pattern = r"\b(2G|3G|4G)\b"
    result = re.findall(pattern, row[3])
   # filename = "2G.xlsx"
    return bool(result)

def filter_3g(row):
    pattern = r"\b(3G|4G)\b"
    result = re.findall(pattern, row[3])
    return bool(result)

def filter_4g(row):
    return "4G" in row[3]

def filter_nomobile(row):
    return row[3] == "-"








@bot.message_handler(commands=['2g'])
def handle_2g_command(message):
    log_user_data_from_message(message)
    bot.send_message(message.chat.id, 'Загружаю данные')
    filter_and_send_data(message, filter_2g, "2G")

@bot.message_handler(commands=['3g'])
def handle_3g_command(message):
    log_user_data_from_message(message)
    bot.send_message(message.chat.id, 'Загружаю данные')
    filter_and_send_data(message, filter_3g, "3G")

@bot.message_handler(commands=['4g'])
def handle_4g_command(message):
    bot.send_message(message.chat.id, 'Загружаю данные')
    filter_and_send_data(message, filter_4g, "4G")

@bot.message_handler(commands=['nomobile'])
def handle_nomobile_command(message):
    bot.send_message(message.chat.id, 'Загружаю данные')
    filter_and_send_data(message, filter_nomobile, "nomobile")


@bot.message_handler(commands=['jt'])
def handle_jt_command(message):
    log_user_data_from_message(message)

    url = f"https://fantastic-engine.vercel.app/"
    bot.send_message(message.chat.id, "😁")
    time.sleep(3)
    bot.send_message(message.chat.id, "🤭", reply_markup=webAppKeyboard_jt(url))






def load_fp_data():
    # Загружаем данные из Google Sheets
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='ФАП!A1:M55').execute()
    rows = result.get('values', [])

    fp_data = []
    fp_headers = ['Адрес', 'Тип подключения', 'Скорость', 'Контакты', 'Дата подписания']

    for row in rows:
        # Выгружаем непустые строки и столбцы B, C, D, F, G, H, I
        if any(row) and "Исключение"  not in row:  # Проверяем, что строка не пустая и не содержит "Исключ"
            filled_row = [row[i] if i < len(row) else '' for i in [1, 2, 3, 5, 6, 7, 8]]
            fp_data.append(filled_row)

    return fp_data, fp_headers

def load_aggoics_data():
    # Загружаем данные из Google Sheets
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='АгГОиЧС!A1:P250').execute()
    rows = result.get('values', [])

    aggoics_data = []
    aggoics_headers = ['Муниципальное образование', 'Наименование населенного пункта', 'Адрес', 'Тип подключения', 'Наименование учреждения', 'Скорость']

    for row in rows:
        # Выгружаем непустые строки и столбцы D, E, F, G, I, J
        if any(row):  # Проверяем, что строка не пустая
            filled_row = [row[i] if i < len(row) else '' for i in [3, 4, 5, 6, 8, 9]]
            aggoics_data.append(filled_row)

    return aggoics_data, aggoics_headers

def load_otpusk_data():
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='otpusk!A1:F30').execute()
    rows = result.get('values', [])
    return rows





def get_employees_on_vacation(otpusk_data, days_ahead=3):
    today = datetime.today().date()
    future_vacation_start = today + timedelta(days=days_ahead)
    employees_on_vacation = []
    employees_starting_vacation_soon = []

    for row_idx, row in enumerate(otpusk_data):
        if row_idx == 0:  # пропустить заголовки таблицы
            continue
        if len(row) >= 5:
            try:
                start_date = datetime.strptime(row[3], "%d.%m.%Y").date()
                end_date = datetime.strptime(row[4], "%d.%m.%Y").date()

                if start_date <= today <= end_date:
                    employees_on_vacation.append(row)

                if today < start_date <= future_vacation_start:
                    employees_starting_vacation_soon.append(row)

            except ValueError:
                pass  # игнорировать строки с неправильным форматом даты

    return employees_on_vacation, employees_starting_vacation_soon







def create_pie_chart(yes_count, no_count, filename):
    labels = ['Есть', 'Нет']
    sizes = [yes_count, no_count]
    colors = ['#2ecc71', '#e74c3c']

    # Создайте объект figure с заданными размерами (ширина, высота) в дюймах
    plt.figure(figsize=(2, 2))  # Здесь 2.5 дюйма - это ширина и высота диаграммы

    plt.pie(sizes, labels=labels, colors=colors, autopct='%1.1f%%', startangle=90)
    plt.axis('equal')
    plt.savefig(filename, dpi=400,
                bbox_inches='tight')  # Установите разрешение (DPI) изображения и параметр bbox_inches
    plt.clf()


def create_bar_chart(data, filename):
    labels, yes_values, no_values = zip(*data)

    total_values = [yes + no for yes, no in zip(yes_values, no_values)]
    yes_percentages = [yes / total * 100 if total != 0 else 0 for yes, total in zip(yes_values, total_values)]
    no_percentages = [no / total * 100 if total != 0 else 0 for no, total in zip(no_values, total_values)]

    labels = labels[1:]
    yes_percentages = yes_percentages[1:]
    no_percentages = no_percentages[1:]

    y = np.arange(len(labels))
    width = 0.6
    colors = ['#2ecc71', '#e74c3c']

    fig, ax = plt.subplots(figsize=(12, 16), dpi=300)  # Устанавливаем размер и DPI изображения
    rects1 = ax.barh(y, yes_percentages, label='Процент подключенных услуг', color=colors[0], align='center')
    rects2 = ax.barh(y, no_percentages, label='Процент не подключенных услуг', left=yes_percentages, color=colors[1],
                     align='center')

    ax.set_title('Подключение к ТОРКНД в Красноярском крае')
    ax.set_yticks(y)
    ax.set_yticklabels(labels)
    ax.legend()

    xmin = 0
    xmax = 100
    ax.set_xlim([xmin, xmax])

    def autolabel(rects, labels):
        for rect, label in zip(rects, labels):
            width = rect.get_width()
            ax.annotate('{:.1f}%'.format(label),
                        xy=(width, rect.get_y() + rect.get_height() / 2),
                        xytext=(3, 0),
                        textcoords="offset points",
                        ha='left', va='center')

    autolabel(rects1, yes_percentages)
    # autolabel(rects2, no_percentages)

    plt.tight_layout()
    plt.savefig(filename)
    plt.close()


data = [
    ('Район 1', 5, 10),
    ('Район 2', 10, 15),
    ('Район 3', 20, 5),
    ('Район 4', 30, 25),
    ('Район 5', 50, 20),
]

create_bar_chart(data, 'output.png')


def search_szofed_values(column_4_value):
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='szofed!A1:M2412').execute()
    rows = result.get('values', [])

    found_values = [row for row in rows if column_4_value.lower() == row[0].lower()]

    return found_values


def search_espd_values(query):
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='espd!A1:AL1466').execute()
    rows = result.get('values', [])
    found_values = [row for row in rows if query.lower() == row[0].lower()]
    return found_values


def search_szoreg_values(query):
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='szoreg!A1:Q1700').execute()
    rows = result.get('values', [])

    found_values = [row for row in rows if query.lower() == row[0].lower()]

    return found_values

def found_mszu_mo(query):
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()
    result = sheet.values().get(spreadsheetId=SPREADSHEET_ID, range='МСЗУ-ОМСУ (тест)!A1:P3200').execute()
    rows = result.get('values', [])

    found_values = [row for row in rows if query.lower() == row[0].lower()]

    return found_values



# def send_district_info(message, district):


# found_values_s = search_values(district)

# response = 'Наименование | Население | Сотовая связь | Интернет | Программа | Таксофон\n'
# response += '-' * 71 + '\n'
#  for row in found_values_s:
# if len(row) >= 19:  # Проверьте, что длина списка row больше или равна максимальному индексу, который вы используете
# response += f"{row[18]} | {row[2]} | {row[3]} | {row[9]} | {row[11]} | {row[12]}\n"


#  messages = split_message(response)
# for msg in messages:
# if msg:  # Проверьте, что сообщение не пустое
#  bot.send_message(message.chat.id, msg, parse_mode='HTML')

# @bot.message_handler(func=lambda message: any(re.match(f"/{district}", message.text) for district in districts))
# def handle_district_command(message):
#  command = message.text[1:]  # Удалите символ '/' из команды
# send_district_info(message, command)

@bot.message_handler(commands=['start'])
def handle_start(message):
    user_first_name = message.from_user.first_name
    bot.send_message(message.chat.id,f'Привет, {user_first_name}!\nЯ бот который может поделиться с тобой информацией о связи в Красноярском крае. Для этого введи название населенного пункта или муниципального образования (например "Курагино" или "Абанский")\nЧтобы узнать информацию о сотовой связи, выбери /2g /3g или /4g. Чтобы получить информацию о населенных пунктах без сотовой связи жми /nomobile \n\n'
        'Для получения списка ФАП из контракта с ПАО "Ростелеком" нажми /fp\n'
        'Для получения списка точек Аг.ГОиЧС из контракта с ПАО "Ростелеком" нажми /ago\n\n'
        'Чтобы узнать о подключении к ТОРКНД, введи сообщение "тор" и наименование муниципального образования. '
        'Например, "тор Енисейский".\n'
        'Если нужна статистика по всему краю, жми /knd_kraj\n\n'
        'Чтобы узнать, кто сегодня в отпуске, жми /otpusk\n\n'
        'Если остались вопросы, пиши @rejoller.')


def preprocess_rows(rows):
    preprocessed_rows = []
    for row in rows:
        if len(row) > 1:
            lemmatized_keywords = {token.lemma_ for token in nlp(row[1].lower()) if not token.is_stop and not token.is_punct}
            preprocessed_rows.append((row, lemmatized_keywords))
    return preprocessed_rows






import spacy

nlp = spacy.load("ru_core_news_sm")

def check_mszu_column_b(user_message):
    # Создание сервиса для доступа к API Google Sheets
    service = build('sheets', 'v4', credentials=creds)

    # Указание диапазона ячеек в таблице "МСЗУ"
    range_name = 'МСЗУ!A1:P150'

    # Получение данных из указанного диапазона ячеек
    result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=range_name).execute()
    values = result.get('values', [])
    preprocessed_values = preprocess_rows(values)

    # Обработка сообщения пользователя с помощью spacy
    user_doc = nlp(user_message.lower())

    # Получение лемматизированных ключевых слов, исключая стоп-слова и пунктуацию
    user_keywords = {token.lemma_ for token in user_doc if not token.is_stop and not token.is_punct}

    # Поиск точных совпадений с ключевыми словами пользователя в столбце B
    matching_rows = [row for row, lemmatized_keywords in preprocessed_values if user_keywords == lemmatized_keywords]


    # Если нет точных совпадений, ищем строки, содержащие хотя бы часть ключевых слов из сообщения пользователя
    if not matching_rows:
        matching_rows = [row for row, lemmatized_keywords in preprocessed_values if user_keywords.intersection(lemmatized_keywords)]


    return matching_rows if len(matching_rows) > 0 else None








import nltk

def ngrams(sequence, n):
    return list(nltk.ngrams(sequence, n))

def jaccard_similarity(a, b):
    a_set = set(a)
    b_set = set(b)
    return len(a_set.intersection(b_set)) / len(a_set.union(b_set))

def preprocess_rows_2(rows):
    preprocessed_rows = []
    for row in rows:
        if len(row) > 1:  # Проверка наличия данных в столбце B
            b_column_value = row[1]  # Индекс 1 соответствует столбцу B
            row_doc = nlp(b_column_value.lower())
            row_keywords = {}
            for i, token in enumerate(row_doc):
                if not token.is_stop and not token.is_punct and len(token) > 2:
                    weight = 2 if i + 1 < len(row_doc) and row_doc[i + 1].text.lower() in ["сельсовет", "район", "округ", "муниципальный округ"] else 1
                    row_keywords[token.lemma_] = weight
            preprocessed_rows.append((row, row_keywords))
    return preprocessed_rows

def weighted_keyword_match(user_keywords, row_keywords):
    user_keyword_set = set(user_keywords.keys())
    row_keyword_set = set(row_keywords.keys())

    intersection = user_keyword_set.intersection(row_keyword_set)
    union = user_keyword_set.union(row_keyword_set)

    if not union:
        return 0

    weighted_intersection_sum = sum([user_keywords.get(k, 0) * row_keywords.get(k, 0) for k in intersection])
    weighted_union_sum = sum([user_keywords.get(k, 0) for k in union]) + sum([row_keywords.get(k, 0) for k in union]) - weighted_intersection_sum

    return weighted_intersection_sum / weighted_union_sum

def check_mszu_mo(user_message):
    user_doc = nlp(user_message.lower())
    user_keywords = {
        token.lemma_: 2 if token.text.lower() in ["сельсовет", "район", "округ", "муниципальный округ"] else 1
        for token in user_doc
        if not token.is_stop and not token.is_punct and len(token) > 2
    }

    # Создание сервиса для доступа к API Google Sheets
    service = build('sheets', 'v4', credentials=creds)

    index_range_name = 'mszuindex!A1:C500'
    main_range_name = 'МСЗУ-ОМСУ (тест)!A1:T3200'

    index_result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=index_range_name).execute()
    main_result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=main_range_name).execute()

    index_values = index_result.get('values', [])
    main_values = main_result.get('values', [])

    preprocessed_index_values = preprocess_rows_2(index_values)

    matching_rows = []
    for index_row, index_keywords in preprocessed_index_values:
        if weighted_keyword_match(user_keywords, index_keywords) >= 0.75:
            index_c_value = index_row[2]  # Получаем значение из столбца C таблицы mszuindex
            # Находим соответствующие строки в таблице "МСЗУ-ОМСУ (тест)" путем сравнения столбца J с index_c_value
            matched_rows = [row for row in main_values if row[9] == index_c_value]
            matching_rows.extend(matched_rows)

    return matching_rows

def preprocess_rows_3(rows):
    preprocessed_rows = []
    for row in rows:
        if len(row) > 5:  # Проверка наличия данных в столбце F
            f_column_value = row[5]  # Индекс 5 соответствует столбцу F
            row_doc = nlp(f_column_value.lower())
            row_keywords = {}
            for i, token in enumerate(row_doc):
                if not token.is_stop and not token.is_punct and len(token) > 2:
                    weight = 2 if i + 1 < len(row_doc) and row_doc[i + 1].text.lower() in ["район", "го", "мо", "округ", "муниципальный округ"] else 1
                    row_keywords[token.lemma_] = weight
            preprocessed_rows.append((row, row_keywords))
    return preprocessed_rows

def weighted_keyword_match(user_keywords, row_keywords):
    user_keyword_set = set(user_keywords.keys())
    row_keyword_set = set(row_keywords.keys())

    intersection = user_keyword_set.intersection(row_keyword_set)
    union = user_keyword_set.union(row_keyword_set)

    if not union:
        return 0

    weighted_intersection_sum = sum([user_keywords.get(k, 0) * row_keywords.get(k, 0) for k in intersection])
    weighted_union_sum = sum([user_keywords.get(k, 0) for k in union]) + sum([row_keywords.get(k, 0) for k in union]) - weighted_intersection_sum

    return weighted_intersection_sum / weighted_union_sum


def check_mszu_mo_2(user_message):
    user_doc = nlp(user_message.lower())
    user_keywords = {
        token.lemma_: 2 if token.text.lower() in ["сельсовет", "район", "округ", "муниципальный округ"] else 1
        for token in user_doc
        if not token.is_stop and not token.is_punct and len(token) > 2
    }

    # Создание сервиса для доступа к API Google Sheets
    service = build('sheets', 'v4', credentials=creds)

    index_range_name = 'mszuindex!A1:G500'
    main_range_name = 'МСЗУ-ОМСУ (тест)!A1:T3200'

    index_result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=index_range_name).execute()
    main_result = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=main_range_name).execute()

    index_values = index_result.get('values', [])
    main_values = main_result.get('values', [])

    preprocessed_index_values = preprocess_rows_3(index_values)

    matching_rows = []
    for index_row, index_keywords in preprocessed_index_values:
        if weighted_keyword_match(user_keywords, index_keywords) >= 0.75:
            index_g_value = index_row[6]  # Получаем значение из столбца G таблицы mszuindex
            # Находим соответствующие строки в таблице "МСЗУ-ОМСУ (тест)" путем сравнения столбца K с index_g_value
            matched_rows = [row for row in main_values if row[10] == index_g_value]
            matching_rows.extend(matched_rows)

    return matching_rows













@bot.message_handler(content_types=['text'])
def handle_text(message):
    print(f"Received message: {message.text}")
    log_user_data_from_message(message)
    if message.text.lower().startswith("тор "):
        handle_tor_message(message)
        return  # добавлено условие
    if message.text.lower() == "кнд край":
        handle_knd_kraj_message(message)
        return
    if message.text.lower().startswith("мсзу "):  # добавлено условие для МСЗУ
        user_first_name = message.from_user.first_name
        bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌')
        handle_mszu_message(message)
        return
    if message.text.lower().startswith("2мсзу "):  # добавлено условие для МСЗУ
        user_first_name = message.from_user.first_name
        bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌 Загружаю информацию по МСЗУ муниципальных образований ')
        handle_2mszu_message(message)
        return

    found_values_a, found_values_s = search_values(message.text)

    found_mszu_values = check_mszu_column_b(message.text)



    if not found_values_a and not found_values_s and not found_mszu_values:
        bot.send_message(message.chat.id, 'Не удалось найти информацию по данному запросу')
        return






   # found_values_a, found_values_s = search_values(message.text)

  #  if not found_values_a and not found_values_s:
      #  bot.send_message(message.chat.id, 'Не удалось найти информацию по данному запросу')
      #  return

    # Если соответствие найдено в столбце A
    if found_values_a:
        #bot.send_message(message.chat.id, f'Секундочку, {user_first_name} Загружаю информацию по вашему запросу ')
        found_values = found_values_a

        if len(found_values) == 1:
            response = f'{found_values[0][1]}:\nЧисленность населения (2010 г.): {found_values[0][2]} чел.\nЧисленность населения (2020 г.): {found_values[0][5]} чел.\nСотовая связь: {found_values[0][3]}\nИнтернет: {get_value(found_values[0], 9)}\nПрограммы: {get_value(found_values[0], 11)}\nКоличество таксофонов: {get_value(found_values[0], 12)}'

            messages = split_message(response)

            for msg in messages:
                bot.send_message(message.chat.id, msg)
            latitude = found_values[0][7]  # Широта находится в столбце H таблицы goroda2.0
            longitude = found_values[0][8]  # Долгота находится в столбце I таблицы goroda2.0

            # Отправляем карту с отмеченной точкой на координатах населенного пункта
            bot.send_location(message.chat.id, latitude, longitude)
            szofed_values = search_szofed_values(found_values[0][4])
            espd_values = search_espd_values(found_values[0][4])
            szoreg_values = search_szoreg_values(found_values[0][4])
            inline_keyboard = types.InlineKeyboardMarkup(row_width=3)
            if szofed_values or espd_values or szoreg_values:

                if szofed_values:
                    szofed_response = 'В указанном населенном пункте рамках федерального проекта в период с 2019 по 2021 год были подключены следующие СЗО:\n\n'
                    for i, row in enumerate(szofed_values, 1):
                        szofed_response += f'\n{i}. {row[8]} {row[9]} по адресу {row[4]}\nТип подключения (Узел связи): {row[10]}\nПропускная способность {row[11]} Мб/сек\nДата подключения:{row[12]}.\n'

                    callback_data = json.dumps({"type": "additional_info", "chat_id": message.chat.id})
                    additional_info_storage[message.chat.id] = szofed_response
                    button_additional_info = types.InlineKeyboardButton("СЗО", callback_data=callback_data)
                    inline_keyboard.add(button_additional_info)

                if espd_values:
                    espd_response = 'Точки подключения к ЕСПД в указанном населенном пункте:\n\n'
                    for i, row in enumerate(espd_values, 1):
                        espd_response += f'\n{i}. {row[12]} по адресу: {row[8]},\nТип подключения: {row[9]},\nСкорость: {row[13]}\nКонтакты ответственного сотрудника:{row[18]}.\n'

                    callback_data = json.dumps({"type": "espd_info", "chat_id": message.chat.id})
                    espd_info_storage[message.chat.id] = espd_response
                    button_espd_info = types.InlineKeyboardButton("ЕСПД", callback_data=callback_data)
                    inline_keyboard.add(button_espd_info)

                # szoreg_values = search_szoreg_values(found_values[0][4])
                if szoreg_values:
                    szoreg_response = 'СЗО в указанном населенном пункте, которым предоставляются услуги за счет средств краевого бюджета:\n\n'
                    for i, row in enumerate(szoreg_values, 1):
                        szoreg_response += f'\n{i}. {row[8]} по адресу {row[5]} \nТип подключения: {row[6]}\nПропускная способность {row[9]}.\n'

                    callback_data = json.dumps({"type": "szoreg_info", "chat_id": message.chat.id})
                    szoreg_info_storage[message.chat.id] = szoreg_response
                    button_szoreg_info = types.InlineKeyboardButton("СЗО (региональный ГК)",callback_data=callback_data)
                    inline_keyboard.add(button_szoreg_info)

                bot.send_message(message.chat.id, "Для получения дополнительной информации нажмите на одну из кнопок ниже:", reply_markup=inline_keyboard)

            bot.callback_query_handler(lambda query: json.loads(query.data)["type"] == "additional_info")(handle_additional_info)
            bot.callback_query_handler(lambda query: json.loads(query.data)["type"] == "espd_info")(handle_espd_info)
            bot.callback_query_handler(lambda query: json.loads(query.data)["type"] == "szoreg_info")(handle_szoreg_info)
        # Если найдено более одного значения

        if len(found_values) > 1:
            # Отправить сообщение со всеми значениями из столбца с индексом 1 и 2
            values = [(get_value(row, 1), get_value(row, 2)) for row in found_values]
            values_with_numbers = [f"{i + 1}. {value[0]}" for i, value in enumerate(values)]
            msg = '\n'.join(values_with_numbers)

            # Разбиваем длинный ответ на части
            messages = split_message(f'Найдено несколько населенных пунктов с таким названием. \n\n{msg}')

            # Отправляем разбитые сообщения
            for msg in messages:
                bot.send_message(message.chat.id, msg)
            # latitude = found_values[0][7]  # Широта находится в столбце H таблицы goroda2.0
            # longitude = found_values[0][8]  # Долгота находится в столбце I таблицы goroda2.0

            # Отправляем карту с отмеченной точкой на координатах населенного пункта
            # bot.send_location(message.chat.id, latitude, longitude)
            # Добавить клавиатуру с порядковыми номерами найденных значений
            buttons = [str(i + 1) for i in range(len(found_values))]
            buttons.append("Отмена")
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
            keyboard.add(*buttons)
            bot.send_message(message.chat.id, 'Выберите номер необходимого населенного пункта:', reply_markup=keyboard)
            bot.register_next_step_handler(message, lambda x: handle_choice(x, found_values, keyboard))

    # Если соответствие найдено в столбце S
    if found_values_s:
        found_values = found_values_s
        data = [
            [row[17], row[2], row[3], row[9], row[11], row[12]]
            for row in found_values_s
        ]
        for row in found_values_s:
            if len(row) >= 20:
                data.append([row[17], row[2], row[3], row[9], row[11], row[12]])
            else:
                print(f"Skipping row due to insufficient elements: {row}")


        if len(found_values) > 0:
            response = 'Наименование | Население | Сотовая связь | Интернет | Программа | Таксофон\n'
            response += '-' * 71 + '\n'
            print(found_values_s)
            for row in found_values:
                response += f"{row[17]} | {row[2]} | {row[3]} | {row[9]} | {row[11]} | {row[12]}\n"

            excel_file = create_excel_file_2(headers, data)
            user_first_name = message.from_user.first_name
            bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌')
            time.sleep(2)
            # messages = split_message_table(headers, data)
            # for msg in messages:
            # if msg:  # Проверьте, что сообщение не пустое
            # bot.send_message(message.chat.id, msg, parse_mode='HTML')
            file_name = found_values[0][18] if found_values else "table"

            with create_excel_file_2(headers, data) as file:
                file.name = f'{file_name}.xlsx'
                bot.send_document(message.chat.id, file)

    if found_mszu_values:
        if len(found_mszu_values) == 1:
            response = f"\nНаименование услуги: \n{found_mszu_values[0][2]}\n№ в Рег.перечне (17-р): {found_mszu_values[0][0]}\nОтвет. РОИВ в рег.перечне: {found_mszu_values[0][3]}\nОтв. в плане: {found_mszu_values[0][6]}\nЕСНСИ: {found_mszu_values[0][8]}\nЕПГУ: {found_mszu_values[0][10]}"
            bot.send_message(message.chat.id, response)

            # Создание inline кнопки с ссылкой
            inline_keyboard = types.InlineKeyboardMarkup()
            url_button = types.InlineKeyboardButton(text="Адрес ИФЗ", url=found_mszu_values[0][13])
            inline_keyboard.add(url_button)

            # Отправка сообщения с inline кнопкой
            bot.send_message(message.chat.id, "Нажмите на кнопку ниже для перехода по ссылке:", reply_markup=inline_keyboard)

        elif len(found_mszu_values) > 1:
            values_with_numbers = [f"{i + 1}. {value[2]}" for i, value in enumerate(found_mszu_values)]
            msg = '\n'.join(values_with_numbers)
            messages = split_message(f'Найдено несколько значений:\n\n{msg}')
            for msg in messages:
                bot.send_message(message.chat.id, msg)

            buttons = [str(i + 1) for i in range(len(found_mszu_values))]
            buttons.append("Отмена")
            keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=3)
            keyboard.add(*buttons)
            bot.send_message(message.chat.id, 'Выберите номер необходимого значения:', reply_markup=keyboard)
            bot.register_next_step_handler(message, lambda x: handle_mszu_choice(x, found_mszu_values, keyboard))

def handle_mszu_message(message):
    print(f"Handling MSZU message: {message.text}")  #
    found_values = check_mszu_mo(message.text)
    if found_values:
        response = ""
        for i, row in enumerate(found_values):
            response += f"{i + 1}. Наименование ОМСУ из ЕСНСИ: {row[10]}\n" \
                        f"№ в Рег.перечне (17-р) : {row[0]}\n" \
                        f"Наименование услуги: {row[2]}\n" \
                        f"Ответ. РОИВ в рег.перечне: {row[3]}\n\n"

        # Разбиваем ответ на части, используя функцию split_message
        response_parts = split_message(response)

        # Отправляем каждую часть ответа по отдельности
        for part in response_parts:
            bot.send_message(message.chat.id, part)
    else:
        bot.send_message(message.chat.id, "Не удалось найти информацию. Попробуйте уточнить ваш запрос.")



def handle_2mszu_message(message):
    print(f"Handling MSZU message: {message.text}")
    found_values = check_mszu_mo_2(message.text)
    if found_values:
        message_storage[message.chat.id] = message.text
        unique_values = list(set([row[12] for row in found_values]))
        unique_values.sort()

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)

        numbered_list = ""
        for i, value in enumerate(unique_values):
            markup.add(types.KeyboardButton(str(i + 1)))  # Добавляем цифры вместо значений
            numbered_list += f"{i + 1}. {value}\n"

        msg = bot.send_message(message.chat.id, f"Выберите необходимое учреждение:\n\n{numbered_list}", reply_markup=markup)
        bot.register_next_step_handler(msg, handle_unique_value_selection)
    else:
        bot.send_message(message.chat.id, "Не удалось найти информацию. Попробуйте уточнить ваш запрос.")




def handle_unique_value_selection(message):
    user_message = message_storage[message.chat.id]
    selected_number = int(message.text)  # Получаем выбранный номер
    found_values = check_mszu_mo_2(user_message)
    unique_values = list(set([row[12] for row in found_values]))
    unique_values.sort()
    selected_value = unique_values[selected_number - 1]  # Получаем значение, соответствующее выбранному номеру

    selected_rows = [row for row in found_values if row[12] == selected_value]
    response = ""
    for i, row in enumerate(selected_rows):
        response += f"{i + 1}. Наименование ОМСУ из ЕСНСИ: {row[10]}\n" \
                    f"№ в Рег.перечне (17-р) : {row[0]}\n" \
                    f"Наименование услуги: {row[2]}\n" \
                    f"Ответ. РОИВ в рег.перечне: {row[3]}\n\n"

    response_parts = split_message(response)

    for part in response_parts:
        if part.strip():
            bot.send_message(message.chat.id, part)









user_messages = {}

def handle_tor_message(message):

    query = message.text[4:]  # Получаем часть сообщения после "тор "
    service = build('sheets', 'v4', credentials=creds)

    # Получаем заголовок таблицы
    header_result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID, range='nst!H1:X1'
    ).execute()
    column_headers = header_result.get('values', [])[0]

    # Поиск совпадений в столбце A на листе nst
    nst_result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID, range='nst!A1:X150'
    ).execute()
    nst_rows = nst_result.get('values', [])
    nst_matches = [
        row for row in nst_rows if row and normalize_text_v2(query) == normalize_text_v2(row[0])
    ]

    if nst_matches:
        all_responses = []
        total_yes_count = 0
        total_no_count = 0

        for row in nst_matches:
            institution_name = row[5]
            response_parts = [f"{institution_name}:\n\n"]
            yes_count = 0
            no_count = 0

            for i in range(7, 24):
                column_header = column_headers[i - 7]
                cell_value = row[i]

                if cell_value == "Есть":
                    cell_value = f"🟢{column_header}\n"
                    yes_count += 1
                elif cell_value == "Нет":
                    cell_value = f"🔴{column_header}\n"
                    no_count += 1

                response_parts.append(cell_value)

            response = " ".join(response_parts)
            all_responses.append(response)
            total_yes_count += yes_count
            total_no_count += no_count

        # Create and save the pie chart for the total counts
        pie_chart_filename = "pie_chart_total.png"
        create_pie_chart(total_yes_count, total_no_count, pie_chart_filename)

        # Combine all responses into one message
        combined_response = "\n\n".join(all_responses)
        messages = split_message(combined_response)

        # Send the pie chart
        bot.send_message(message.chat.id, 'Диаграмма подключения к ТОРКНД')
        # Создаем inline кнопку "подробно"
        markup = types.InlineKeyboardMarkup()
        detailed_button = types.InlineKeyboardButton("подробно", callback_data='{"type": "additional_info_tor"}')
        markup.add(detailed_button)

        # Отправляем изображение с кнопкой
        with open(pie_chart_filename, 'rb') as chart_file:
            callback_data = json.dumps({
                "type": "additional_info_tor",
                "user_id": message.from_user.id
            })
            detailed_button = types.InlineKeyboardButton("подробно", callback_data=callback_data)
            markup = types.InlineKeyboardMarkup()
            markup.add(detailed_button)

            # Сохраняем сообщения для пользователя в глобальном словаре
            user_messages[message.from_user.id] = messages

            bot.send_photo(message.chat.id, chart_file, reply_markup=markup)

            # Remove the pie chart file after sending
        os.remove(pie_chart_filename)
        user_messages[message.chat.id] = messages
    else:
        bot.send_message(message.chat.id, "Не найдено совпадений для запроса")

def handle_mszu_choice(message, found_mszu_values, keyboard):
    choice = message.text.strip()
    if choice.isdigit():
        index = int(choice) - 1
        if 0 <= index < len(found_mszu_values):
            response = f"\nНаименование услуги: \n{found_mszu_values[index][2]}\n№ в Рег.перечне (17-р): {found_mszu_values[index][0]}\nОтвет. РОИВ в рег.перечне: {found_mszu_values[index][3]}\nОтв. в плане: {found_mszu_values[index][6]}\nЕСНСИ: {found_mszu_values[index][8]}\nЕПГУ: {found_mszu_values[index][10]}"
            bot.send_message(message.chat.id, response, reply_markup=types.ReplyKeyboardRemove())
            inline_keyboard = types.InlineKeyboardMarkup()
            url_button = types.InlineKeyboardButton(text="Адрес ИФЗ", url=found_mszu_values[index][13])
            inline_keyboard.add(url_button)

            # Отправка сообщения с inline кнопкой
            bot.send_message(message.chat.id, "Нажмите на кнопку ниже для перехода по ссылке:", reply_markup=inline_keyboard)
        else:
            bot.send_message(message.chat.id, 'Неверный номер. Попробуйте еще раз или нажмите "Отмена".', reply_markup=keyboard)
            bot.register_next_step_handler(message, lambda x: handle_mszu_choice(x, found_mszu_values, keyboard))
    elif choice.lower() == "отмена":
        bot.send_message(message.chat.id, 'Поиск отменен.', reply_markup=types.ReplyKeyboardRemove())
    else:
        bot.send_message(message.chat.id, 'Введите корректный номер или нажмите "Отмена".', reply_markup=keyboard)
        bot.register_next_step_handler(message, lambda x: handle_mszu_choice(x, found_mszu_values, keyboard))




# Обработчик нажатия кнопки
@bot.callback_query_handler(func=lambda call: json.loads(call.data)["type"] == "additional_info_tor")
def detailed_button_callback(call):
    user_first_name = call.from_user.first_name
    bot.send_message(call.message.chat.id, f'Секундочку, {user_first_name}😌 Загружаю статистику для тебя')
    time.sleep(3)

    # Получаем messages из глобального словаря
    callback_data = json.loads(call.data)
    user_id = callback_data["user_id"]
    messages = user_messages.get(user_id, [])

    for msg in messages:
        bot.send_message(call.message.chat.id, msg)
    bot.answer_callback_query(call.id)
    time.sleep(2)
    bot.send_message(call.message.chat.id, 'Введите свой следующий запрос')




def handle_knd_kraj_message(message):
    service = build('sheets', 'v4', credentials=creds)

    # Получаем все строки таблицы nst
    nst_result = service.spreadsheets().values().get(
        spreadsheetId=SPREADSHEET_ID, range='nst!A1:X150'
    ).execute()
    nst_rows = nst_result.get('values', [])

    # Подсчет количества значений "Есть" и "Нет" для каждой строки в столбце A
    counter = {}
    for row in nst_rows:
        key = row[0]
        yes_count = sum([1 for value in row[7:24] if value == "Есть"])
        no_count = sum([1 for value in row[7:24] if value == "Нет"])

        if key in counter:
            counter[key] = (key, yes_count + counter[key][1], no_count + counter[key][2])
        else:
            counter[key] = (key, yes_count, no_count)

    # Создание столбчатой диаграммы
    bar_chart_filename = 'bar_chart.png'
    create_bar_chart(list(counter.values()), bar_chart_filename)

    # Отправка столбчатой диаграммы
    with open(bar_chart_filename, 'rb') as chart_file:
        bot.send_photo(message.chat.id, chart_file)

    # Удаление файла диаграммы после отправки
    os.remove(bar_chart_filename)

COLUMNS_TO_EXPORT = [1, 2, 3, 6, 7]

# Заголовки таблицы
TABLE_HEADERS = ["Наименование", "Население", "Сотовая связь", "Интернет", "Таксофон"]


def handler_fp_message(message, fp_data, fp_headers):

    time.sleep(1)
    try:
        data = [
            [row[0], row[1], row[2], row[4], row[5]]
            for row in fp_data[1:]
        ]
        headers = ['Муниципальное образование', 'Наименование населенного пункта', 'Адрес', 'Наименование учреждения', 'Скорость']

        excel_file = create_excel_file(headers, data)
        user_first_name = message.from_user.first_name
        #bot.send_message(message.chat.id, f'Секундочку, {user_first_name}😌')
        #time.sleep(2)
        bot.send_message(message.chat.id, 'Перечень ФАП из госконтракта:')
        file_name = 'ФАП'

        with BytesIO(excel_file) as file:
            file.name = f'{file_name}.xlsx'
            bot.send_document(message.chat.id, file)

    except Exception as e:
        logging.exception(e)
        bot.send_message(message.chat.id, "Произошла ошибка. Попробуйте еще раз.")

def handler_aggoics_message(message, aggoics_data, aggoics_headers):

    time.sleep(1)
    try:
        data = [
            [row[0], row[1], row[2], row[3], row[4], row[5]]
            for row in aggoics_data[1:]
        ]

        excel_file = create_excel_file(aggoics_headers, data)



        bot.send_message(message.chat.id, 'Информация из таблицы АгГОиЧС:')
        file_name = 'АгГОиЧС'

        with BytesIO(excel_file) as file:
            file.name = f'{file_name}.xlsx'
            bot.send_document(message.chat.id, file)

    except Exception as e:
        logging.exception(e)
        bot.send_message(message.chat.id, "Произошла ошибка. Попробуйте еще раз.")


def handler_otpusk_message(message, employees_on_vacation):
    if len(employees_on_vacation) > 0:
        response = "Сотрудники, которые сегодня в отпуске:\n\n"
        for employee in employees_on_vacation:
            response += f"{employee[0]} ({employee[1]})\n"
        time.sleep(2)
        bot.send_message(message.chat.id, response)
    else:
        time.sleep(2)
        bot.send_message(message.chat.id, "Сегодня никто не в отпуске.")





def handle_choice(message, found_values, keyboard):
    szoreg_response = ""
    espd_response = ""
    if message.text == "Отмена":
        bot.send_message(message.chat.id, 'Поиск отменен.', reply_markup=types.ReplyKeyboardRemove())
        return
    try:
        index = int(message.text)
        if index <= 0 or index > len(found_values):
            raise ValueError

        response = f'{get_value(found_values[index - 1], 1)}:\nЧисленность населения (2010 г): {get_value(found_values[index - 1], 2)} чел.\nЧисленность населения (2010 г): {get_value(found_values[index - 1], 5)} чел.\nСотовая связь: {get_value(found_values[index - 1], 3)}\nИнтернет: {get_value(found_values[index - 1], 9)}\nПрограммы: {get_value(found_values[index - 1], 11)}\nКоличество таксофонов: {get_value(found_values[index - 1], 12)}'

        messages = split_message(response)

        for msg in messages:
            bot.send_message(message.chat.id, msg)
        latitude = found_values[index - 1][7]  # Широта находится в столбце H таблицы goroda2.0
        longitude = found_values[index - 1][8]  # Долгота находится в столбце I таблицы goroda2.0

        # Отправляем карту с отмеченной точкой на координатах населенного пункта
        bot.send_location(message.chat.id, latitude, longitude)
        inline_keyboard = types.InlineKeyboardMarkup()

        szofed_values = search_szofed_values(found_values[index - 1][4])
        if szofed_values:
            szofed_response = 'В указанном населенном пункте рамках федерального проекта в период с 2019 по 2021 год были подключены следующие СЗО:\n\n'
            for i, row in enumerate(szofed_values, 1):
                szofed_response += f'\n{i}. {row[8]} {row[9]} по адресу {row[4]}\nТип подключения (Узел связи): {row[10]}\nПропускная способность {row[11]} Мб/сек\nДата подключения:{row[12]}.\n'

            callback_data = json.dumps({"type": "additional_info", "chat_id": message.chat.id})
            additional_info_storage[message.chat.id] = szofed_response
            button_additional_info = types.InlineKeyboardButton("СЗО", callback_data=callback_data)
            inline_keyboard.add(button_additional_info)

        espd_values = search_espd_values(found_values[index - 1][4])
        if espd_values:
            espd_response = 'Точки подключения к ЕСПД в указанном населенном пункте:\n\n'
            for i, row in enumerate(espd_values, 1):
                espd_response += f'\n{i}. {row[12]} по адресу: {row[8]},\nТип подключения: {row[9]},\nСкорость: {row[13]}\nКонтакты ответственного сотрудника:{row[18]}.\n'

            callback_data = json.dumps({"type": "espd_info", "chat_id": message.chat.id})
            espd_info_storage[message.chat.id] = espd_response
            button_espd_info = types.InlineKeyboardButton("ЕСПД", callback_data=callback_data)
            inline_keyboard.add(button_espd_info)

        szoreg_values = search_szoreg_values(found_values[index - 1][4])
        if szoreg_values:
            szoreg_response = 'СЗО в указанном населенном пункте, которым предоставляются услуги за счет средств краевого бюджета:\n\n'
            for i, row in enumerate(szoreg_values, 1):
                szoreg_response += f'\n{i}. {row[8]} по адресу {row[5]} \nТип подключения: {row[6]}\nПропускная способность {row[9]}.\n'

            callback_data = json.dumps({"type": "szoreg_info", "chat_id": message.chat.id})
            szoreg_info_storage[message.chat.id] = szoreg_response
            button_szoreg_info = types.InlineKeyboardButton("СЗО (региональный ГК)", callback_data=callback_data)

        if szofed_values or espd_values or szoreg_values:
            bot.send_message(message.chat.id, "Для получения дополнительной информации нажмите на кнопку ниже", reply_markup=inline_keyboard)
        #bot.send_message(message.chat.id, "Для получения дополнительной информации нажмите на кнопку ниже", reply_markup=inline_keyboard)
        bot.callback_query_handler(lambda query: json.loads(query.data)["type"] == "additional_info")(handle_additional_info)
        bot.callback_query_handler(lambda query: json.loads(query.data)["type"] == "espd_info")(handle_espd_info)
        bot.callback_query_handler(lambda query: json.loads(query.data)["type"] == "szoreg_info")(handle_szoreg_info)

        return
    except ValueError:
        bot.send_message(message.chat.id, 'Неверный номер населенного пункта. Попробуйте еще раз.')


def handle_additional_info(query):
    chat_id = json.loads(query.data)["chat_id"]
    if chat_id in additional_info_storage:
        response = additional_info_storage[chat_id]
        messages = split_message(response)
        for message_group in messages:
            msg = ''.join(message_group)
            if msg.strip():  # Проверка, что сообщение не пустое
                bot.send_message(chat_id, msg)

        bot.answer_callback_query(query.id)
    else:
        bot.answer_callback_query(query.id, "Дополнительная информация недоступна.")


def handle_espd_info(query):
    chat_id = json.loads(query.data)["chat_id"]
    if chat_id in espd_info_storage:
        response = espd_info_storage[chat_id]
        messages = split_message(response)
        for message_group in messages:
            msg = ''.join(message_group)
            if msg.strip():  # Проверка, что сообщение не пустое
                bot.send_message(chat_id, msg)

        bot.answer_callback_query(query.id)
    else:
        bot.answer_callback_query(query.id, "Информация из таблицы ЕСПД недоступна.")


def handle_szoreg_info(query):
    chat_id = json.loads(query.data)["chat_id"]
    if chat_id in szoreg_info_storage:
        response = szoreg_info_storage[chat_id]
        messages = split_message(response)
        for message_group in messages:
            msg = ''.join(message_group)
            if msg.strip():  # Проверка, что сообщение не пустое
                bot.send_message(chat_id, msg)

        bot.answer_callback_query(query.id)
    else:
        bot.answer_callback_query(query.id, "Информация из таблицы СЗО (региональный контракт) недоступна.")






if __name__ == "__main__":
    bot.polling(none_stop=True, timeout=100)  # Установите значение таймаута, например, на 50 секунд
    while True:
        try:
            bot.polling(none_stop=True, interval=0)
        except ReadTimeoutError:
            print("Ошибка таймаута, повторное подключение через 5 секунд")
            time.sleep(5)  # Задержка перед повторным подключением
